"""
=============================================================================
ANALIZADOR DE REPORTES SEMANALES DE OBRA (RSO)
Proyecto: Línea de Transferencia Auca 123 a Auca 51
=============================================================================

INSTRUCCIONES DE USO:
  1. Instalar dependencias:  pip install pandas openpyxl matplotlib
  2. Ejecutar:               python analizador_rso.py
  3. Seleccionar uno o varios archivos .xlsx desde el diálogo de archivos.
  4. El informe se guarda en  informe_<nombre_archivo>.txt
     La gráfica se guarda en  curva_<nombre_archivo>.png
     Los datos quedan en      rso_historico.db  (SQLite)

ESTRUCTURA ESPERADA DEL EXCEL:
  - Hoja "RDO": encabezados en fila 13, datos de proyecto en filas 3-9.
  - Hoja "CURVA": encabezados en fila 8, datos diarios desde fila 9.
=============================================================================
"""

import os
import sys
import psycopg2
import datetime
import warnings
warnings.filterwarnings("ignore")

import pandas as pd
import openpyxl
import matplotlib
matplotlib.use("Agg")          # backend sin pantalla (para guardar PNG)
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

try:
    import tkinter as tk
    from tkinter import filedialog
    HAS_TK = True
except ImportError:
    HAS_TK = False


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────
DB_PATH = "rso_historico.db"


# ─────────────────────────────────────────────────────────────────────────────
# 1. SELECCIÓN DE ARCHIVOS
# ─────────────────────────────────────────────────────────────────────────────
def seleccionar_archivos():
    """Abre un diálogo para seleccionar uno o varios archivos .xlsx."""
    if HAS_TK:
        root = tk.Tk()
        root.withdraw()
        archivos = filedialog.askopenfilenames(
            title="Seleccionar Reportes Semanales de Obra (.xlsx)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        root.destroy()
        return list(archivos)
    else:
        # Fallback: recibir rutas por argumento de línea de comandos
        if len(sys.argv) > 1:
            return sys.argv[1:]
        print("ERROR: tkinter no disponible. Pase los archivos como argumentos.")
        return []


# ─────────────────────────────────────────────────────────────────────────────
# 2. LECTURA Y PROCESAMIENTO DE LA HOJA RDO
# ─────────────────────────────────────────────────────────────────────────────
def leer_hoja_rdo(wb):
    """
    Lee la hoja RDO del workbook y retorna:
      - meta:  dict con datos de cabecera del proyecto
      - df:    DataFrame con todas las actividades (consolidadas por código)
      - totales: dict con valores totales del proyecto
    """
    ws = wb["RDO"]

    # ── 2a. Localizar la fila de encabezados ──────────────────────────────
    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Código Preciario":
                header_row = cell.row
                break
        if header_row:
            break

    if header_row is None:
        raise ValueError("No se encontró la fila de encabezados en la hoja RDO.")

    # ── 2b. Construir mapa columna-letra → nombre de campo ────────────────
    # La fila de encabezados (ej. 13) tiene los nombres principales.
    # La siguiente fila (14) tiene sub-encabezados como Inicio/Fin.
    header_map = {}   # col_index (0-based) → nombre
    row_13 = list(ws.iter_rows(min_row=header_row,
                               max_row=header_row, values_only=True))[0]
    row_14 = list(ws.iter_rows(min_row=header_row + 1,
                               max_row=header_row + 1, values_only=True))[0]

    for idx, val in enumerate(row_13):
        if val is not None:
            name = str(val).strip()
            # Combine with sub-header if applicable
            sub = row_14[idx] if idx < len(row_14) and row_14[idx] is not None else ""
            sub = str(sub).strip() if sub else ""
            if sub:
                name = f"{name} {sub}"
            header_map[idx] = name
        elif idx < len(row_14) and row_14[idx] is not None:
            # Sub-header without main header (e.g., Inicio / Fin)
            sub = str(row_14[idx]).strip()
            # Find last known parent header
            parent = header_map.get(idx - 1, "")
            # Avoid duplicating already-named cols
            full_name = f"{parent} {sub}".strip() if parent else sub
            header_map[idx] = full_name

    # Fix known cols where sub-rows complete the name
    # Map by position for critical columns
    col_names_fixed = {
        "Código Preciario": None,
        "Nombre Actividad": None,
        "% Buget": None,
        "% EAC": None,
        "Cantidad Budget": None,
        "Cantidad EAC": None,
        "UM": None,
        "P.U": None,
        "Costo Budget": None,
        "Costo EAC": None,
        "AC Anterior": None,
        "AC Semanal": None,
        "AC Actual": None,
        "EV Anterior": None,
        "EV Semanal": None,
        "EV Actual": None,
        "Earn Value (EAC)": None,
        "PV Costo Planificado": None,
        "Previsto Inicio": None,
        "Previsto Fin": None,
        "Real Inicio": None,
        "Real Fin": None,
        "Cantidad Anterior": None,
        "Cantidad Semanal": None,
        "Cantidad Acumulada": None,
        "% Avance Acumulado Anterior": None,
        "% Avance Semanal": None,
        "% AvanceActual Acumulado": None,
        "% Avance Previsto Acumulado": None,
    }

    # Remap based on actual column letters found in the file
    # Row 13 is header_row; columns (0-based) from openpyxl row
    all_rows_vals = list(ws.iter_rows(min_row=header_row,
                                      max_row=header_row + 2, values_only=True))
    r1 = all_rows_vals[0]   # main headers
    r2 = all_rows_vals[1]   # sub-headers

    # Build final column list combining r1 + r2
    col_list = []
    prev_main = None
    for i in range(len(r1)):
        main = str(r1[i]).strip() if r1[i] is not None else None
        sub  = str(r2[i]).strip() if i < len(r2) and r2[i] is not None else None
        if main:
            prev_main = main
        if sub and prev_main:
            col_list.append(f"{prev_main} {sub}")
        elif main:
            col_list.append(main)
        elif sub:
            col_list.append(sub)
        else:
            col_list.append(f"_col{i}")

    # Normalize duplicates
    seen = {}
    final_cols = []
    for c in col_list:
        if c in seen:
            seen[c] += 1
            final_cols.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            final_cols.append(c)

    # ── 2c. Leer datos de actividades ─────────────────────────────────────
    data_rows = []
    # Skip header rows (header_row, header_row+1, header_row+2 are headers/blanks)
    data_start = header_row + 3   # row 16 in file (0-based: row 15)
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if all(v is None for v in row):
            continue
        # Pad/trim to col_list length
        row_padded = list(row) + [None] * max(0, len(final_cols) - len(row))
        data_rows.append(row_padded[:len(final_cols)])

    df = pd.DataFrame(data_rows, columns=final_cols)

    # ── 2d. Limpiar y tipar columnas ──────────────────────────────────────
    # Remove system rows (B column contains '#' markers)
    if "#FILA_TITULO" in df.columns or "_col1" in df.columns:
        marker_col = [c for c in df.columns if "col1" in c or c == "_col1"]
        if marker_col:
            df = df[~df[marker_col[0]].astype(str).str.startswith("#", na=False)]

    # Remove rows where Nombre Actividad is None
    nombre_col = next((c for c in df.columns if "Nombre Actividad" in c), None)
    codigo_col = next((c for c in df.columns if "Código Preciario" in c), None)
    if nombre_col:
        df = df[df[nombre_col].notna()].copy()

    df.reset_index(drop=True, inplace=True)

    # Numeric columns
    num_cols = [
        c for c in df.columns
        if any(kw in c for kw in [
            "Costo", "AC ", "EV ", "PV", "Cantidad", "% ", "P.U",
            "Avance", "Budget", "EAC",
        ])
    ]
    for c in num_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # Date columns
    date_cols = [c for c in df.columns if any(kw in c for kw in ["Inicio", "Fin", "Previsto", "Real"])]
    for c in date_cols:
        df[c] = pd.to_datetime(df[c], errors="coerce")

    # ── 2e. Extraer totales (primera fila de datos = CONSTRUCCIÓN) ────────
    # The total row is the first row with non-null Costo Budget
    costo_budget_col = next((c for c in df.columns if "Costo Budget" in c), None)
    totales = {}
    if costo_budget_col:
        total_row = df[df[costo_budget_col].notna()].iloc[0] if not df[df[costo_budget_col].notna()].empty else None
        if total_row is not None:
            totales = total_row.to_dict()

    # ── 2f. Filtrar solo actividades con Código Preciario ─────────────────
    if codigo_col:
        df_actividades = df[df[codigo_col].notna() & (df[codigo_col] != "")].copy()
    else:
        df_actividades = df.copy()

    # ── 2g. Consolidar rubros duplicados (mismo Código Preciario) ─────────
    if codigo_col and not df_actividades.empty:
        df_actividades = _consolidar_por_codigo(df_actividades, codigo_col, nombre_col)

    # ── 2h. Extraer metadata del proyecto ─────────────────────────────────
    meta = _extraer_metadata(ws)

    return meta, df_actividades, totales, df


def _consolidar_por_codigo(df, codigo_col, nombre_col):
    """
    Agrupa filas con el mismo Código Preciario:
      - Columnas numéricas → suma
      - P.U → promedio ponderado (o primer valor si no hay cantidad)
      - Columnas no numéricas → primer valor
    """
    num_cols = df.select_dtypes(include="number").columns.tolist()
    # P.U no debe sumarse
    pu_col = next((c for c in df.columns if c == "P.U"), None)
    if pu_col and pu_col in num_cols:
        num_cols.remove(pu_col)

    str_cols = [c for c in df.columns if c not in num_cols and c != codigo_col]

    agg_dict = {c: "sum" for c in num_cols}
    agg_dict.update({c: "first" for c in str_cols})
    if pu_col:
        agg_dict[pu_col] = "first"

    return df.groupby(codigo_col, as_index=False).agg(agg_dict)


def _extraer_metadata(ws):
    """Extrae los campos de cabecera de la hoja RDO (filas 3-9)."""
    meta = {}
    all_vals = {r: [] for r in range(1, 15)}
    for row in ws.iter_rows(min_row=1, max_row=14, values_only=True):
        pass  # just iterate to warm up

    rows = list(ws.iter_rows(min_row=1, max_row=14, values_only=True))

    # Row 3 → proyecto
    meta["proyecto"] = rows[2][2] if len(rows) > 2 else None
    # Row 6 → fecha reporte (col C = index 2)
    meta["fecha_reporte"] = rows[5][2] if len(rows) > 5 else None
    # Row 7 → documento (col I = index 8)
    doc_raw = rows[6][8] if len(rows) > 6 else None
    if doc_raw:
        meta["documento_no"] = str(doc_raw).replace("DOCUMENTO NO: ", "").strip()
    else:
        meta["documento_no"] = None

    # Row 9 → SPI col AI (index 32), CPI col AJ (index 33)
    row9 = rows[8] if len(rows) > 8 else []
    meta["spi"] = float(row9[32]) if len(row9) > 32 and row9[32] is not None and not isinstance(row9[32], str) else None
    meta["cpi"] = float(row9[33]) if len(row9) > 33 and row9[33] is not None and not isinstance(row9[33], str) else None
    meta["fecha_inicio_plan"]  = row9[2]  if len(row9) > 2  else None
    meta["fecha_fin_plan"]     = row9[6]  if len(row9) > 6  else None
    meta["fecha_inicio_real"]  = row9[8]  if len(row9) > 8  else None
    meta["fecha_fin_real"]     = row9[23] if len(row9) > 23 else None
    meta["plazo_previsto"]     = row9[26] if len(row9) > 26 else None
    meta["dias_disruptivos"]   = row9[28] if len(row9) > 28 else None
    meta["plazo_ajustado"]     = row9[30] if len(row9) > 30 else None

    # Convert dates
    for fld in ["fecha_reporte", "fecha_inicio_plan", "fecha_fin_plan",
                "fecha_inicio_real"]:
        v = meta.get(fld)
        if isinstance(v, datetime.datetime):
            meta[fld] = v.date()
        elif isinstance(v, str):
            try:
                meta[fld] = datetime.datetime.strptime(v, "%Y-%m-%d").date()
            except Exception:
                pass

    return meta


# ─────────────────────────────────────────────────────────────────────────────
# 3. LECTURA Y PROCESAMIENTO DE LA HOJA CURVA
# ─────────────────────────────────────────────────────────────────────────────
def leer_hoja_curva(wb):
    """
    Lee la hoja CURVA y retorna un DataFrame con columnas:
      Fecha, % Previsto Acumulado, % Real Acumulado
    """
    ws = wb["CURVA"]
    rows = list(ws.iter_rows(values_only=True))

    # Buscar fila de encabezado (contiene "Fecha")
    header_row_idx = None
    for i, row in enumerate(rows):
        if any(str(c).strip() == "Fecha" for c in row if c is not None):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("No se encontró encabezado 'Fecha' en la hoja CURVA.")

    header = rows[header_row_idx]
    # Find column indices
    col_fecha    = next((i for i, v in enumerate(header) if v is not None and str(v).strip() == "Fecha"), None)
    col_previsto = next((i for i, v in enumerate(header) if v is not None and "Previsto" in str(v)), None)
    col_real     = next((i for i, v in enumerate(header) if v is not None and "Real" in str(v)), None)

    if None in (col_fecha, col_previsto, col_real):
        raise ValueError("No se encontraron columnas esperadas en la hoja CURVA.")

    data = []
    for row in rows[header_row_idx + 1:]:
        if row[col_fecha] is None:
            continue
        fecha    = row[col_fecha]
        previsto = row[col_previsto]
        real     = row[col_real]
        if isinstance(fecha, (datetime.datetime, datetime.date)):
            data.append({
                "Fecha": fecha,
                "% Previsto Acumulado": previsto,
                "% Real Acumulado": real
            })

    df = pd.DataFrame(data)
    df["Fecha"] = pd.to_datetime(df["Fecha"])
    df["% Previsto Acumulado"] = pd.to_numeric(df["% Previsto Acumulado"], errors="coerce")
    df["% Real Acumulado"]     = pd.to_numeric(df["% Real Acumulado"], errors="coerce")
    df.dropna(subset=["% Previsto Acumulado", "% Real Acumulado"], how="all", inplace=True)

    # Convert fractions to percentages if values are ≤ 1
    for col in ["% Previsto Acumulado", "% Real Acumulado"]:
        if df[col].max() <= 1.0:
            df[col] = df[col] * 100

    return df


# ─────────────────────────────────────────────────────────────────────────────
# 4. GENERACIÓN DE LA GRÁFICA
# ─────────────────────────────────────────────────────────────────────────────
def generar_grafica(df_curva, nombre_archivo, output_path):
    """
    Genera la gráfica de % Previsto vs % Real acumulado y la guarda como PNG.
    """
    fig, ax = plt.subplots(figsize=(14, 6))

    # Filtrar solo hasta hoy para la curva real
    hoy = pd.Timestamp.today()
    df_plot = df_curva.copy()

    ax.plot(df_plot["Fecha"], df_plot["% Previsto Acumulado"],
            label="% Previsto Acumulado", color="#2196F3", linewidth=2,
            linestyle="--", marker="")
    ax.plot(df_plot["Fecha"], df_plot["% Real Acumulado"],
            label="% Real Acumulado", color="#4CAF50", linewidth=2.5,
            marker="")

    # Fill between curves to highlight gap
    ax.fill_between(df_plot["Fecha"],
                    df_plot["% Previsto Acumulado"],
                    df_plot["% Real Acumulado"],
                    where=(df_plot["% Previsto Acumulado"] > df_plot["% Real Acumulado"]),
                    alpha=0.15, color="red", label="Brecha de atraso")

    ax.fill_between(df_plot["Fecha"],
                    df_plot["% Previsto Acumulado"],
                    df_plot["% Real Acumulado"],
                    where=(df_plot["% Real Acumulado"] >= df_plot["% Previsto Acumulado"]),
                    alpha=0.15, color="green", label="Adelanto")

    # Last real value marker
    ultima_real = df_plot[df_plot["% Real Acumulado"].notna()].iloc[-1]
    ax.scatter(ultima_real["Fecha"], ultima_real["% Real Acumulado"],
               color="#4CAF50", s=80, zorder=5)
    ax.annotate(f'{ultima_real["% Real Acumulado"]:.2f}%',
                xy=(ultima_real["Fecha"], ultima_real["% Real Acumulado"]),
                xytext=(10, 10), textcoords="offset points",
                fontsize=9, color="#2E7D32",
                arrowprops=dict(arrowstyle="->", color="#2E7D32"))

    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b-%y"))
    ax.xaxis.set_major_locator(mdates.MonthLocator())
    plt.xticks(rotation=45)
    ax.set_xlabel("Fecha", fontsize=11)
    ax.set_ylabel("Avance Acumulado (%)", fontsize=11)
    ax.set_title(f"Curva S – % Previsto vs % Real Acumulado\n{nombre_archivo}",
                 fontsize=13, fontweight="bold")
    ax.legend(loc="upper left", fontsize=9)
    ax.grid(True, alpha=0.3)
    ax.set_ylim(0, 105)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.0f}%"))

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  → Gráfica guardada: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# 5. GENERACIÓN DEL INFORME DE TEXTO
# ─────────────────────────────────────────────────────────────────────────────
def _fmt_num(v, decimals=2, prefix=""):
    """Formatea un número con separador de miles."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "N/D"
    try:
        fmt = f"{v:,.{decimals}f}"
        return f"{prefix}{fmt}"
    except Exception:
        return str(v)


def _fmt_date(v):
    if v is None:
        return "No registrada"
    if isinstance(v, (datetime.datetime, datetime.date)):
        return str(v)[:10]
    return str(v)


def generar_informe(nombre_archivo, meta, df_actividades, totales, df_curva, df_full=None):
    """Genera el texto completo del informe."""

    lines = []
    sep = "=" * 72

    # ── Encabezado ────────────────────────────────────────────────────────
    lines.append(sep)
    lines.append("  REPORTE SEMANAL DE OBRA – ANÁLISIS AUTOMÁTICO")
    lines.append(sep)
    lines.append(f"Nombre del archivo analizado: {nombre_archivo}")
    lines.append(f"Fecha del informe: {_fmt_date(meta.get('fecha_reporte'))}")
    lines.append(f"Fecha de procesamiento: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")

    # ── Resumen ejecutivo ─────────────────────────────────────────────────
    lines.append("📋 RESUMEN EJECUTIVO")
    lines.append("-" * 65)
    spi = meta.get("spi")
    cpi = meta.get("cpi")
    try:
        spi = float(spi) if spi is not None else None
    except (ValueError, TypeError):
        spi = None
    try:
        cpi = float(cpi) if cpi is not None else None
    except (ValueError, TypeError):
        cpi = None
    spi_str = f"{spi:.4f}" if spi is not None else "N/D"
    cpi_str = f"{cpi:.4f}" if cpi is not None else "N/D"
    
    # MODIFICADO: Resumen ejecutivo conciso con fecha, SPI y CPI
    lines.append(f"Fecha: {_fmt_date(meta.get('fecha_reporte'))}  |  SPI: {spi_str}  |  CPI: {cpi_str}")
    lines.append("")

    # ── Datos generales ───────────────────────────────────────────────────
    lines.append("📊 DATOS GENERALES DEL PROYECTO")
    lines.append("-" * 65)
    tabla_gen = [
        ("Proyecto",                meta.get("proyecto", "N/D")),
        ("Documento No",            meta.get("documento_no", "N/D")),
        ("Fecha del Reporte",       _fmt_date(meta.get("fecha_reporte"))),
        ("Fecha Inicio Planificada",_fmt_date(meta.get("fecha_inicio_plan"))),
        ("Fecha Fin Planificada",   _fmt_date(meta.get("fecha_fin_plan"))),
        ("Fecha Inicio Real",       _fmt_date(meta.get("fecha_inicio_real"))),
        ("Fecha Fin Real",          _fmt_date(meta.get("fecha_fin_real"))),
        ("Plazo Previsto (días)",   str(meta.get("plazo_previsto", "N/D"))),
        ("Días Disruptivos",        str(meta.get("dias_disruptivos", "N/D"))),
        ("Plazo Ajustado (días)",   str(meta.get("plazo_ajustado", "N/D"))),
        ("SPI",                     spi_str),
        ("CPI",                     cpi_str),
    ]
    col_w = 30
    for concepto, valor in tabla_gen:
        lines.append(f"  {concepto:<{col_w}} {valor}")
    lines.append("")
    lines.append("Interpretación de Índices:")
    if spi is not None:
        if spi < 1:
            lines.append(f"  SPI < 1 ({spi_str}): El proyecto está ATRASADO respecto al cronograma.")
        elif spi == 1:
            lines.append(f"  SPI = 1: El proyecto está EN LÍNEA con el cronograma.")
        else:
            lines.append(f"  SPI > 1 ({spi_str}): El proyecto está ADELANTADO respecto al cronograma.")
    if cpi is not None:
        if cpi < 1:
            lines.append(f"  CPI < 1 ({cpi_str}): El proyecto está POR ENCIMA del presupuesto.")
        elif cpi == 1:
            lines.append(f"  CPI = 1: El proyecto está EN LÍNEA con el presupuesto.")
        else:
            lines.append(f"  CPI > 1 ({cpi_str}): El proyecto está POR DEBAJO del presupuesto (eficiente en costos).")
    lines.append("")

    # ── Análisis financiero EVM ───────────────────────────────────────────
    lines.append("💰 ANÁLISIS FINANCIERO Y DE VALOR GANADO (EVM)")
    lines.append("-" * 65)

    # Extraer valores del dict de totales o del DataFrame
    def get_total(keys):
        for k in keys:
            col = next((c for c in totales.keys() if k in str(c)), None)
            if col and totales.get(col) is not None:
                v = totales[col]
                if isinstance(v, (int, float)) and not pd.isna(v):
                    return float(v)
        return None

    bac  = get_total(["Costo Budget"])
    eac  = get_total(["Costo EAC"])
    ac   = get_total(["AC Actual"])
    ev   = get_total(["EV Actual"])
    pv   = get_total(["PV Costo"])

    tabla_evm = [
        ("Costo Budget (BAC)",             _fmt_num(bac)),
        ("Costo EAC",                      _fmt_num(eac)),
        ("Costo Real Acumulado (AC)",       _fmt_num(ac)),
        ("Valor Ganado Acumulado (EV)",     _fmt_num(ev)),
        ("Valor Planificado Acumulado (PV)",_fmt_num(pv)),
    ]
    for concepto, valor in tabla_evm:
        lines.append(f"  {concepto:<40} {valor}")
    lines.append("")
    lines.append("Variaciones Clave:")
    if ev is not None and ac is not None:
        cv = ev - ac
        cv_status = "Favorable (bajo costo)" if cv >= 0 else "Desfavorable (sobrecosto)"
        lines.append(f"  CV = EV - AC = {cv:,.2f} → {cv_status}")
    if ev is not None and pv is not None:
        sv = ev - pv
        sv_status = "Favorable (adelanto)" if sv >= 0 else "Desfavorable (atraso)"
        lines.append(f"  SV = EV - PV = {sv:,.2f} → {sv_status}")
    if bac is not None and eac is not None:
        ahorro = bac - eac
        lines.append(f"  Proyección de ahorro/sobrecosto al cierre = {ahorro:,.2f}")
    lines.append("")

    # ── Avance físico ─────────────────────────────────────────────────────
    lines.append("📈 ANÁLISIS DE AVANCE FÍSICO")
    lines.append("-" * 65)

    # Tomar último valor de la curva (solo donde real es > 0)
    avance_real_curva = None
    avance_prev_curva = None
    if not df_curva.empty:
        df_real_valido = df_curva[df_curva["% Real Acumulado"] > 0]
        if not df_real_valido.empty:
            avance_real_curva = df_real_valido.iloc[-1]["% Real Acumulado"]
        # Último previsto hasta hoy
        hoy = pd.Timestamp.today()
        df_prev_pasado = df_curva[df_curva["Fecha"] <= hoy]
        if not df_prev_pasado.empty:
            avance_prev_curva = df_prev_pasado.iloc[-1]["% Previsto Acumulado"]

    # Fallback: from totales
    if avance_real_curva is None:
        ar_raw = get_total(["% AvanceActual Acumulado", "% Avance Actual"])
        avance_real_curva = ar_raw * 100 if ar_raw is not None and ar_raw <= 1 else ar_raw
    if avance_prev_curva is None:
        ap_raw = get_total(["% Avance Previsto Acumulado", "% Avance Previsto"])
        avance_prev_curva = ap_raw * 100 if ap_raw is not None and ap_raw <= 1 else ap_raw

    retraso = None
    if avance_prev_curva is not None and avance_real_curva is not None:
        retraso = avance_prev_curva - avance_real_curva

    tabla_avance = [
        ("% Avance Actual Acumulado",   _fmt_num(avance_real_curva) + "%"),
        ("% Avance Previsto Acumulado", _fmt_num(avance_prev_curva) + "%"),
        ("Retraso Físico",              _fmt_num(retraso) + "%"),
    ]
    for concepto, valor in tabla_avance:
        lines.append(f"  {concepto:<35} {valor}")
    lines.append("")
    if avance_real_curva and avance_prev_curva:
        lines.append(
            f"  El proyecto ha completado solo {avance_real_curva:.2f}% del trabajo, "
            f"cuando debería haber completado {avance_prev_curva:.2f}%, "
            f"lo que confirma el atraso señalado por el SPI."
        )
    lines.append("")

    # ── WBS por tramo ─────────────────────────────────────────────────────
    lines.append("🏗️ ESTRUCTURA DEL TRABAJO (WBS) POR TRAMO")
    lines.append("-" * 65)

    # Use df_full (all rows) to find tramo-level rows
    df_wbs = df_full if df_full is not None else df_actividades
    nombre_col = next((c for c in (df_wbs.columns if df_wbs is not None and not df_wbs.empty else [])
                       if "Nombre Actividad" in c), None)
    pct_budget_col = next((c for c in (df_wbs.columns if df_wbs is not None and not df_wbs.empty else [])
                           if "% Buget" in c), None)
    pct_eac_col    = next((c for c in (df_wbs.columns if df_wbs is not None and not df_wbs.empty else [])
                           if "% EAC" in c), None)
    avance_col     = next((c for c in (df_wbs.columns if df_wbs is not None and not df_wbs.empty else [])
                           if "% AvanceActual Acumulado" in c), None)

    tramos_df = None
    if nombre_col is not None and df_wbs is not None:
        mask = df_wbs[nombre_col].astype(str).str.contains(
            r"Línea de Flujo|Línea Aérea|Línea Enterrada", na=False, regex=True)
        # Exclude the top-level "Instalación de Línea de Flujo..." summary row
        mask2 = ~df_wbs[nombre_col].astype(str).str.startswith("Instalación de Línea", na=False)
        tramos_df = df_wbs[mask & mask2].copy()

    if tramos_df is not None and not tramos_df.empty:
        lines.append(f"  Se identificaron {len(tramos_df)} tramos principales en el proyecto.")
        lines.append("")
        lines.append(f"  {'Tramo':<55} {'%Budget':>8}  {'%EAC':>8}  {'%Avance':>8}")
        lines.append("  " + "-" * 85)
        for _, row in tramos_df.iterrows():
            nombre = str(row.get(nombre_col, ""))[:54]
            pb = row.get(pct_budget_col)
            pe = row.get(pct_eac_col)
            pa = row.get(avance_col)
            pb_str = f"{pb*100:.2f}%" if pd.notna(pb) and pb <= 1 else (f"{pb:.2f}%" if pd.notna(pb) else "N/D")
            pe_str = f"{pe*100:.2f}%" if pd.notna(pe) and pe <= 1 else (f"{pe:.2f}%" if pd.notna(pe) else "N/D")
            pa_str = f"{pa*100:.2f}%" if pd.notna(pa) and pa <= 1 else (f"{pa:.2f}%" if pd.notna(pa) else "N/D")
            lines.append(f"  {nombre:<55} {pb_str:>8}  {pe_str:>8}  {pa_str:>8}")
    else:
        lines.append("  No se identificaron tramos tipo 'Línea de Flujo' en los datos.")
    lines.append("")

  
    # ── Conclusión ────────────────────────────────────────────────────────
    lines.append("🧮 CONCLUSIÓN")
    lines.append("-" * 65)
    estado, emoji = _determinar_estado(spi, cpi, retraso)
    lines.append(f"  Estado General: {emoji} {estado}")
    # Eliminado el párrafo explicativo posterior
    lines.append("")
    lines.append(sep)

    return "\n".join(lines)


def _identificar_hallazgos(df, meta):
    """Genera lista de hallazgos automáticos."""
    hallazgos = []
    if df.empty:
        hallazgos.append("No se pudieron analizar actividades (DataFrame vacío).")
        return hallazgos

    avance_col    = next((c for c in df.columns if "% AvanceActual Acumulado" in c), None)
    real_fin_col  = next((c for c in df.columns if "Real Fin" in c), None)
    prev_fin_col  = next((c for c in df.columns if "Previsto Fin" in c), None)
    ac_sem_col    = next((c for c in df.columns if "AC Semanal" in c), None)
    real_ini_col  = next((c for c in df.columns if "Real Inicio" in c), None)
    prev_ini_col  = next((c for c in df.columns if "Previsto Inicio" in c), None)

    hoy = pd.Timestamp.today()

    # Actividades con avance 0 y fecha fin pasada
    if avance_col and prev_fin_col:
        try:
            mask = (
                (pd.to_numeric(df[avance_col], errors="coerce").fillna(0) == 0) &
                (pd.to_datetime(df[prev_fin_col], errors="coerce") < hoy)
            )
            n_zero = mask.sum()
            if n_zero > 0:
                hallazgos.append(
                    f"Atraso Generalizado: {n_zero} actividad(es) con 0% de avance "
                    f"habiendo superado su fecha fin planificada."
                )
        except Exception:
            pass

    # Costos negativos (AC Semanal)
    if ac_sem_col:
        try:
            neg = (pd.to_numeric(df[ac_sem_col], errors="coerce") < 0).sum()
            if neg > 0:
                hallazgos.append(
                    f"Valores Negativos en Costos Semanales: {neg} actividad(es) con "
                    f"AC Semanal negativo (posibles reversiones o ajustes contables)."
                )
        except Exception:
            pass

    # Fechas reales no registradas
    if real_ini_col:
        try:
            sin_inicio = df[real_ini_col].isna().sum()
            if sin_inicio > 0:
                hallazgos.append(
                    f"Fechas Reales No Registradas: {sin_inicio} actividad(es) sin fecha de inicio real."
                )
        except Exception:
            pass

    if real_fin_col:
        try:
            sin_fin = df[real_fin_col].isna().sum()
            if sin_fin > 0:
                hallazgos.append(
                    f"Fechas de Fin Real Ausentes: {sin_fin} actividad(es) sin fecha de fin real."
                )
        except Exception:
            pass

    # Costo EAC vs BAC
    costo_budget_col = next((c for c in df.columns if "Costo Budget" in c), None)
    costo_eac_col    = next((c for c in df.columns if "Costo EAC" in c), None)
    if costo_budget_col and costo_eac_col:
        try:
            bac_total = pd.to_numeric(df[costo_budget_col], errors="coerce").sum()
            eac_total = pd.to_numeric(df[costo_eac_col],    errors="coerce").sum()
            if eac_total < bac_total:
                ahorro = bac_total - eac_total
                hallazgos.append(
                    f"Costo EAC Inferior al BAC: Se proyecta un ahorro de "
                    f"aprox. {ahorro:,.2f} al final del proyecto."
                )
            elif eac_total > bac_total:
                sobrecosto = eac_total - bac_total
                hallazgos.append(
                    f"Sobrecosto Proyectado: EAC supera BAC en {sobrecosto:,.2f}."
                )
        except Exception:
            pass

    # SPI/CPI
    try: spi = float(meta.get("spi")) if meta.get("spi") is not None else None
    except Exception: spi = None
    try: cpi = float(meta.get("cpi")) if meta.get("cpi") is not None else None
    except Exception: cpi = None
    if spi is not None and spi < 0.8:
        hallazgos.append(f"SPI Crítico ({spi:.4f}): El proyecto presenta retraso severo en cronograma.")
    if cpi is not None and cpi > 1.3:
        hallazgos.append(
            f"CPI Alto ({cpi:.4f}): Monitorear que la alta eficiencia de costos "
            f"no esté afectando alcance o calidad."
        )
    if cpi is not None and cpi < 0.9:
        hallazgos.append(f"CPI Crítico ({cpi:.4f}): El proyecto presenta sobrecosto significativo.")

    if not hallazgos:
        hallazgos.append("No se detectaron hallazgos críticos automáticos.")

    return hallazgos


def _generar_recomendaciones(hallazgos, spi, cpi, retraso):
    """Genera recomendaciones basadas en los hallazgos."""
    recs = []
    texto_hallazgos = " ".join(hallazgos)

    if "atraso" in texto_hallazgos.lower() or (spi and spi < 1):
        recs.append("Recuperación de Cronograma: Implementar acciones correctivas "
                    "para acelerar las actividades críticas atrasadas.")
    if "fecha" in texto_hallazgos.lower():
        recs.append("Actualización de Fechas Reales: Completar las fechas reales de "
                    "inicio y fin para mejorar la trazabilidad del proyecto.")
    if "negativo" in texto_hallazgos.lower():
        recs.append("Validación de Costos: Investigar y documentar las razones de "
                    "los valores de AC Semanal negativos.")
    if "0%" in texto_hallazgos or "avance 0" in texto_hallazgos.lower():
        recs.append("Análisis de Causa Raíz: Investigar por qué actividades no han "
                    "iniciado a pesar de superar sus fechas planificadas.")
    if cpi is not None and cpi > 1.3:
        recs.append("Monitoreo de CPI Alto: Auditar que la eficiencia en costos no "
                    "esté ocultando problemas de alcance o calidad.")
    recs.append("Validación de Datos: Verificar periódicamente la exactitud de los "
                "porcentajes de avance y costos reportados.")
    recs.append("Reunión de Revisión Semanal: Mantener reuniones regulares de "
                "seguimiento con el equipo de campo.")
    return recs


def _determinar_estado(spi, cpi, retraso):
    """Determina el estado general del proyecto."""
    spi_ok = spi is not None and spi >= 0.95
    cpi_ok = cpi is not None and cpi >= 0.95
    retraso_ok = retraso is not None and retraso <= 5

    if spi_ok and cpi_ok and retraso_ok:
        return "Verde (En Control)", "🟢"
    elif (spi is not None and spi < 0.75) or (cpi is not None and cpi < 0.75):
        return "Rojo (Estado Crítico)", "🔴"
    else:
        return "Amarillo (Riesgo Moderado – Atraso en Cronograma)", "🟡"


# ─────────────────────────────────────────────────────────────────────────────
# 6. BASE DE DATOS PostgreSQL (Neon)
# ─────────────────────────────────────────────────────────────────────────────

import os
import psycopg2
import datetime
import pandas as pd

DATABASE_URL = os.environ.get("DATABASE_URL")


def inicializar_db():
    """Crea la tabla si no existe."""
    conn = psycopg2.connect(DATABASE_URL)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS reportes (
            id                  SERIAL PRIMARY KEY,
            nombre_archivo      TEXT,
            fecha_reporte       TEXT,
            spi                 REAL,
            cpi                 REAL,
            avance_real         REAL,
            avance_planificado  REAL,
            costo_budget        REAL,
            costo_eac           REAL,
            costo_real          REAL,
            ev                  REAL,
            pv                  REAL,
            fecha_inicio_plan   TEXT,
            fecha_fin_plan      TEXT,
            fecha_inicio_real   TEXT,
            fecha_fin_real      TEXT,
            plazo_previsto      INTEGER,
            dias_disruptivos    INTEGER,
            plazo_ajustado      INTEGER,
            fecha_procesamiento TEXT
        )
    """)

    conn.commit()
    cursor.close()
    conn.close()


def guardar_en_db(nombre_archivo, meta, totales, avance_real, avance_planificado):
    """Inserta o actualiza el registro del reporte en PostgreSQL."""
    inicializar_db()

    def _safe(v):
        if isinstance(v, (datetime.date, datetime.datetime)):
            return str(v)[:10]
        return v

    def get_total_val(keys):
        for k in keys:
            col = next((c for c in totales.keys() if k in str(c)), None)
            if col:
                v = totales.get(col)
                if v is not None and isinstance(v, (int, float)) and not pd.isna(v):
                    return float(v)
        return None

    record = (
        nombre_archivo,
        _safe(meta.get("fecha_reporte")),
        meta.get("spi"),
        meta.get("cpi"),
        avance_real,
        avance_planificado,
        get_total_val(["Costo Budget"]),
        get_total_val(["Costo EAC"]),
        get_total_val(["AC Actual"]),
        get_total_val(["EV Actual"]),
        get_total_val(["PV Costo"]),
        _safe(meta.get("fecha_inicio_plan")),
        _safe(meta.get("fecha_fin_plan")),
        _safe(meta.get("fecha_inicio_real")),
        _safe(meta.get("fecha_fin_real")),
        meta.get("plazo_previsto"),
        meta.get("dias_disruptivos"),
        meta.get("plazo_ajustado"),
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )

    conn = psycopg2.connect(DATABASE_URL)
    cursor = conn.cursor()

    # Verificar si ya existe el registro
    cursor.execute(
        "SELECT id FROM reportes WHERE nombre_archivo = %s AND fecha_reporte = %s",
        (record[0], record[1])
    )
    existing = cursor.fetchone()

    if existing:
        cursor.execute("""
            UPDATE reportes SET
                spi=%s,
                cpi=%s,
                avance_real=%s,
                avance_planificado=%s,
                costo_budget=%s,
                costo_eac=%s,
                costo_real=%s,
                ev=%s,
                pv=%s,
                fecha_inicio_plan=%s,
                fecha_fin_plan=%s,
                fecha_inicio_real=%s,
                fecha_fin_real=%s,
                plazo_previsto=%s,
                dias_disruptivos=%s,
                plazo_ajustado=%s,
                fecha_procesamiento=%s
            WHERE id=%s
        """, record[2:] + (existing[0],))

        print(f"  → DB actualizado (id={existing[0]}): {nombre_archivo}")

    else:
        cursor.execute("""
            INSERT INTO reportes (
                nombre_archivo,
                fecha_reporte,
                spi,
                cpi,
                avance_real,
                avance_planificado,
                costo_budget,
                costo_eac,
                costo_real,
                ev,
                pv,
                fecha_inicio_plan,
                fecha_fin_plan,
                fecha_inicio_real,
                fecha_fin_real,
                plazo_previsto,
                dias_disruptivos,
                plazo_ajustado,
                fecha_procesamiento
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, record)

        new_id = cursor.fetchone()[0]
        print(f"  → DB insertado (id={new_id}): {nombre_archivo}")

    conn.commit()
    cursor.close()
    conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# 7. PROCESO PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────
def procesar_archivo(ruta_archivo):
    """Procesa un único archivo Excel RSO."""
    nombre_archivo = os.path.basename(ruta_archivo)
    nombre_base    = os.path.splitext(nombre_archivo)[0]

    print(f"\n{'='*60}")
    print(f"  Procesando: {nombre_archivo}")
    print(f"{'='*60}")

    # ── Cargar workbook ───────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
    except Exception as e:
        print(f"  ERROR al abrir el archivo: {e}")
        return

    # ── Leer hojas ────────────────────────────────────────────────────────
    try:
        print("  Leyendo hoja RDO...")
        meta, df_actividades, totales, df_full = leer_hoja_rdo(wb)
    except Exception as e:
        print(f"  ERROR al procesar hoja RDO: {e}")
        import traceback; traceback.print_exc()
        return

    try:
        print("  Leyendo hoja CURVA...")
        df_curva = leer_hoja_curva(wb)
    except Exception as e:
        print(f"  ADVERTENCIA hoja CURVA: {e}")
        df_curva = pd.DataFrame(columns=["Fecha","% Previsto Acumulado","% Real Acumulado"])

    # ── Obtener avances para DB ───────────────────────────────────────────
    avance_real = None
    avance_plan = None
    if not df_curva.empty:
        df_r = df_curva[df_curva["% Real Acumulado"] > 0]
        if not df_r.empty:
            avance_real = float(df_r.iloc[-1]["% Real Acumulado"])
        df_p = df_curva[df_curva["Fecha"] <= pd.Timestamp.today()]
        if not df_p.empty:
            avance_plan = float(df_p.iloc[-1]["% Previsto Acumulado"])

    # ── Gráfica ───────────────────────────────────────────────────────────
    grafica_path = f"curva_{nombre_base}.png"
    if not df_curva.empty:
        print("  Generando gráfica...")
        generar_grafica(df_curva, nombre_archivo, grafica_path)
    else:
        print("  Sin datos de curva – gráfica omitida.")

    # ── Informe de texto ──────────────────────────────────────────────────
    print("  Generando informe...")
    informe = generar_informe(nombre_archivo, meta, df_actividades, totales, df_curva, df_full=df_full)
    informe_path = f"informe_{nombre_base}.txt"
    with open(informe_path, "w", encoding="utf-8") as f:
        f.write(informe)
    print(f"  → Informe guardado: {informe_path}")

    # ── Base de datos ─────────────────────────────────────────────────────
    print("  Guardando en base de datos SQLite...")
    guardar_en_db(nombre_archivo, meta, totales, avance_real, avance_plan)

    # ── Mostrar informe en consola ────────────────────────────────────────
    print("\n" + informe)

    return informe_path, grafica_path


def milyy():
    """Punto de entrada principal."""
    print("=" * 60)
    print("  ANALIZADOR RSO – Proyecto Auca 123 a Auca 51")
    print("=" * 60)

    archivos = seleccionar_archivos()

    if not archivos:
        print("No se seleccionaron archivos. Saliendo.")
        return

    resultados = []
    for archivo in archivos:
        if not os.path.isfile(archivo):
            print(f"  Archivo no encontrado: {archivo}")
            continue
        resultado = procesar_archivo(archivo)
        if resultado:
            resultados.append(resultado)

    print(f"\n{'='*60}")
    print(f"  Procesamiento completado. {len(resultados)} archivo(s) analizados.")
    print(f"  Base de datos: {os.path.abspath(DB_PATH)}")
    print("=" * 60)


if __name__ == "__main__":
    milyy()